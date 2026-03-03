#!/usr/bin/env python3
"""ETF 邮件提醒脚本。

功能概览：
1) 抓取 ETF（CSPX.GB / CSNDX.CH / SMH.GB / IWY.US / COWZ.NL）行情（当前价、当日最高、当日最低）
2) 抓取 CNN Fear & Greed 指标
3) 根据抓取结果生成 PASS / NOK 邮件并通过 SMTP 发送
4) 支持 --once 单次执行和按北京时间工作日定时执行
"""

import argparse
import base64
import configparser
from decimal import ROUND_HALF_UP, Decimal
import gzip
import html as html_lib
import json
import re
import smtplib
import sys
import time
import zlib
from dataclasses import dataclass
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

# 明确使用北京时间，避免服务器本地时区差异造成定时偏移
BEIJING_TZ = ZoneInfo("Asia/Shanghai")
# 统一请求头，降低被目标站点拒绝的概率
USER_AGENT = "Mozilla/5.0 (ETF-Reminder/1.0; +https://github.com/aky56/homework)"
# 单次 HTTP 请求超时（秒）
REQUEST_TIMEOUT = 10
# 失败重试次数（总尝试次数 = RETRY_COUNT + 1）
RETRY_COUNT = 2
BOC_WHPJ_URL = "https://www.boc.cn/sourcedb/whpj/"
BOC_TARGET_CURRENCY = "美元"
PINGAN_GOLD_DETAIL_URL = (
    "https://rmb.pingan.com.cn/bron/ibank/cust/bron-ibank-pd/gold/product/acct/getPopGoldDetail.do?prdCode=GOLD_SHARE"
)
JM_GOLD_FEAR_GREED_URL = "https://cdn.jmbullion.com/fearandgreed/fearandgreed.json"
CNN_DATAVIZ_API_BASE = "https://production.dataviz.cnn.io"
CNN_MARKET_INDEX_SYMBOLS = "DJII-USA,SP500-CME,COMP-USA"
JPM_118UC_URL = (
    "https://am.jpmorgan.com/lu/en/asset-management/per/products/jpm-us-technology-a-acc-usd-lu0210536867"
)
JPM_118UC_DATA_URL = (
    "https://am.jpmorgan.com/FundsMarketingHandler/product-data?cusip=LU0210536867&country=lu&role=per&language=en"
)
BLACKROCK_134UU_URL = "https://www.blackrock.com/cn/products/230010/bgf-world-technology-fund-a2-usd"
BLACKROCK_077UC_URL = (
    "https://www.blackrock.com/cn/products/270404/bgf-global-enhanced-equity-yield-fund-a6-usd"
)
SC_172UC_PAGE_URL = (
    "https://www.sc.com/cn/investment/funds/fund-details/?isin=LU2708338459-%E7%BE%8E%E5%85%83&fund-type=qdmf"
)
SC_GRAPHQL_URL = "https://www.sc.com/cn/graphql/"
SC_172UC_ISIN = "LU2708338459"
SC_172UC_CURRENCY = "USD"
AB_FUNDS_API_BASE_URL = "https://webapi.alliancebernstein.com"
AB_089UU_PAGE_URL = (
    "https://www.abfunds.com.hk/hk/zh-hk/investor/funds/equities/ab-low-volatility-equity.ad.LU0965508806.html"
)
AB_123UU_PAGE_URL = (
    "https://www.abfunds.com.hk/hk/zh-hk/investor/funds/equities/ab-low-volatility-total-return-equity.ad.LU1934455277.html"
)
AB_089UU_SHARE_CLASS_ID = "LU0965508806"
AB_123UU_SHARE_CLASS_ID = "LU1934455277"
LOG_XLSX_FILENAME = "usser_log.xlsx"
LOG_XLSX_PATH_OVERRIDE: Optional[Path] = None
LOG_XLSX_SHEET1_NAME = "Sheet1"
LOG_XLSX_SHEET2_NAME = "Sheet2"
LOG_XLSX_SHEET3_NAME = "Sheet3"
LOG_XLSX_HEADERS = [
    "时间",
    "汇率",
    "Fear & Greed",
    "S&P 500",
    "S&P 500 涨跌",
    "NASDAQ",
    "NASDAQ 涨跌",
    "黄金价格",
    "黄金情绪指数",
    "118UC",
    "134UU",
    "077UC",
    "172UC",
    "089UU",
    "123UU",
]
LOG_XLSX_ETF_HEADERS = [
    "时间",
    "CSPX.GB.CUR",
    "CSPX.GB.LOW",
    "CSPX.GB.HIG",
    "CSNDX.CH.CUR",
    "CSNDX.CH.LOW",
    "CSNDX.CH.HIG",
    "IWY.US.CUR",
    "IWY.US.LOW",
    "IWY.US.HIG",
    "SMH.GB.CUR",
    "SMH.GB.LOW",
    "SMH.GB.HIG",
    "COWZ.NL.CUR",
    "COWZ.NL.LOW",
    "COWZ.NL.HIG",
]
LOG_XLSX_ETF_SYMBOL_ORDER = ["CSPX.GB", "CSNDX.CH", "IWY.US", "SMH.GB", "COWZ.NL"]

# 浏览器风格请求头：降低被反爬误判风险
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
}

# 目标 ETF 与可用数据源候选（按顺序依次尝试）
ETF_SYMBOL_CANDIDATES: Dict[str, List[Tuple[str, str]]] = {
    "CSPX.GB": [
        ("yahoo", "CSPX.L"),
        ("stooq", "cspx.gb"),
    ],
    "CSNDX.CH": [
        ("yahoo", "CSNDX.SW"),
        ("stooq", "csndx.ch"),
    ],
    "SMH.GB": [
        ("yahoo", "SMH.L"),
        ("stooq", "smh.gb"),
    ],
    "IWY.US": [
        ("yahoo", "IWY"),
        ("stooq", "iwy.us"),
    ],
    "COWZ.NL": [
        ("yahoo", "COWZ.AS"),
        ("stooq", "cowz.nl"),
    ],
}


@dataclass
class PriceResult:
    """ETF 行情抓取结果。"""

    symbol: str
    current: Optional[float]
    day_high: Optional[float]
    day_low: Optional[float]
    source: Optional[str]
    error: Optional[str]


@dataclass
class FearGreedResult:
    """Fear & Greed 指标抓取结果。"""

    value: Optional[float]
    rating: Optional[str]
    error: Optional[str]


@dataclass
class CnnMarketResult:
    """CNN 页面市场摘要结果（S&P 500 / NASDAQ）。"""

    sp500_value: Optional[str]
    sp500_change: Optional[str]
    nasdaq_value: Optional[str]
    nasdaq_change: Optional[str]
    error: Optional[str]


@dataclass
class IndexSnapshot:
    """指数快照（点位 + 涨跌幅）。"""

    value: Optional[str]
    change: Optional[str]
    error: Optional[str]


@dataclass
class SmtpConfig:
    """SMTP 配置。"""

    host: str
    port: int
    username: str
    password: str
    from_email: str
    use_tls: bool
    use_ssl: bool


@dataclass
class ExchangeRateResult:
    """中行汇率抓取结果（美元现汇卖出价）。"""

    rate: Optional[float]
    publish_date: Optional[str]
    publish_time: Optional[str]
    error: Optional[str]


@dataclass
class GoldPriceResult:
    """黄金价格抓取结果。"""

    value: Optional[float]
    error: Optional[str]


@dataclass
class GoldSentimentResult:
    """黄金情绪指数抓取结果。"""

    value: Optional[float]
    error: Optional[str]


@dataclass
class FundPriceResult:
    """基金价格抓取结果。"""

    value: Optional[float]
    error: Optional[str]


class _BOCPriceTableParser(HTMLParser):
    """解析中国银行外汇牌价页，提取指定币种所在行。"""

    def __init__(self, target_currency: str) -> None:
        super().__init__()
        self.target_currency = target_currency
        self.in_price_table = False
        self.table_depth = 0
        self.in_tr = False
        self.in_td = False
        self.current_cell: List[str] = []
        self.current_row: List[str] = []
        self.target_row: Optional[List[str]] = None

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        attr_map = dict(attrs)
        if tag == "table" and attr_map.get("id") == "priceTable":
            self.in_price_table = True
            self.table_depth = 1
            return

        if not self.in_price_table:
            return

        if tag == "table":
            self.table_depth += 1
            return
        if tag == "tr":
            self.in_tr = True
            self.current_row = []
            return
        if tag == "td" and self.in_tr:
            self.in_td = True
            self.current_cell = []

    def handle_data(self, data: str) -> None:
        if self.in_price_table and self.in_td:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if not self.in_price_table:
            return

        if tag == "td" and self.in_td:
            self.current_row.append("".join(self.current_cell).strip())
            self.in_td = False
            self.current_cell = []
            return

        if tag == "tr" and self.in_tr:
            if self.current_row and self.current_row[0] == self.target_currency:
                self.target_row = self.current_row[:]
            self.in_tr = False
            self.current_row = []
            return

        if tag == "table":
            self.table_depth -= 1
            if self.table_depth <= 0:
                self.in_price_table = False


def log(message: str) -> None:
    """统一日志输出格式（带北京时间时间戳）。"""

    now = datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S %Z")
    print(f"[{now}] {message}")


def _decode_body(raw: bytes, content_encoding: str) -> str:
    """按响应压缩格式解码为文本。"""

    encoding = (content_encoding or "").lower()
    data = raw
    if "gzip" in encoding:
        data = gzip.decompress(raw)
    elif "deflate" in encoding:
        try:
            data = zlib.decompress(raw, -zlib.MAX_WBITS)
        except zlib.error:
            data = zlib.decompress(raw)
    return data.decode("utf-8", errors="replace")


def _http_get_decoded_text(url: str, extra_headers: Optional[Dict[str, str]] = None) -> str:
    """发送 GET 请求并自动处理 gzip/deflate 压缩与重试。"""

    headers = dict(DEFAULT_HEADERS)
    if extra_headers:
        headers.update(extra_headers)

    last_error: Optional[Exception] = None
    for attempt in range(RETRY_COUNT + 1):
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=REQUEST_TIMEOUT) as response:
                raw = response.read()
                return _decode_body(raw, response.headers.get("Content-Encoding", ""))
        except (HTTPError, URLError, TimeoutError, UnicodeDecodeError, zlib.error, OSError) as exc:
            last_error = exc
            if attempt < RETRY_COUNT:
                delay = 2 ** attempt
                log(f"请求失败，{delay}s 后重试: {exc}")
                time.sleep(delay)
    raise RuntimeError(f"请求失败: {last_error}")


def http_get_json(url: str, extra_headers: Optional[Dict[str, str]] = None) -> dict:
    """发送 GET 请求并解析 JSON。

    - 使用统一 User-Agent
    - timeout=10
    - 失败后指数退避重试（1s, 2s）
    """

    text = _http_get_decoded_text(url, extra_headers=extra_headers)
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"JSON 解析失败: {exc}") from exc


def http_post_json(url: str, payload: dict, extra_headers: Optional[Dict[str, str]] = None) -> dict:
    """发送 POST 请求并解析 JSON（带重试）。"""

    headers = dict(DEFAULT_HEADERS)
    headers["Content-Type"] = "application/json"
    if extra_headers:
        headers.update(extra_headers)

    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    last_error: Optional[Exception] = None
    for attempt in range(RETRY_COUNT + 1):
        try:
            req = Request(url, data=body, headers=headers, method="POST")
            with urlopen(req, timeout=REQUEST_TIMEOUT) as response:
                raw = response.read()
                text = _decode_body(raw, response.headers.get("Content-Encoding", ""))
                return json.loads(text)
        except (HTTPError, URLError, TimeoutError, UnicodeDecodeError, zlib.error, OSError, json.JSONDecodeError) as exc:
            last_error = exc
            if attempt < RETRY_COUNT:
                delay = 2 ** attempt
                log(f"请求失败，{delay}s 后重试: {exc}")
                time.sleep(delay)
    raise RuntimeError(f"请求失败: {last_error}")


def http_get_text(url: str) -> str:
    """发送 GET 请求并返回文本。

    逻辑同 `http_get_json`，只是不做 JSON 解析。
    """

    return _http_get_decoded_text(url)


def fetch_yahoo_price(symbol: str) -> Tuple[float, float, float]:
    """从 Yahoo 抓取行情，返回 (当前价, 当日最高, 当日最低)。"""

    query = urlencode({"range": "1d", "interval": "5m"})
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?{query}"
    data = http_get_json(url)
    result = data.get("chart", {}).get("result")
    if not result:
        raise RuntimeError(f"Yahoo 返回无数据: {data.get('chart', {}).get('error')}")

    info = result[0]
    meta = info.get("meta", {})
    quote = info.get("indicators", {}).get("quote", [{}])[0]
    current = meta.get("regularMarketPrice")
    highs = [h for h in (quote.get("high") or []) if h is not None]
    lows = [l for l in (quote.get("low") or []) if l is not None]

    # 若关键字段缺失则抛错，由上层兜底并切换备用数据源
    if current is None or not highs or not lows:
        raise RuntimeError("Yahoo 数据字段缺失")

    return float(current), float(max(highs)), float(min(lows))


def fetch_stooq_price(symbol: str) -> Tuple[float, float, float]:
    """从 Stooq 抓取 CSV 行情，返回 (当前价, 当日最高, 当日最低)。"""

    url = f"https://stooq.com/q/l/?s={symbol}&f=sd2t2ohlcv&h&e=csv"
    text = http_get_text(url)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        raise RuntimeError("Stooq 返回数据为空")

    header = [h.strip().lower() for h in lines[0].split(",")]
    values = [v.strip() for v in lines[1].split(",")]
    row = dict(zip(header, values))
    if row.get("close") in (None, "N/D"):
        raise RuntimeError("Stooq close 字段缺失")

    current = float(row["close"])
    day_high = float(row["high"])
    day_low = float(row["low"])
    return current, day_high, day_low


def fetch_price(symbol: str) -> PriceResult:
    """抓取指定 ETF 行情。

    按候选源顺序逐个尝试；如果全部失败，返回带错误信息的结果，
    但不抛异常，确保流程可继续生成 NOK 邮件。
    """

    errors: List[str] = []
    for source, source_symbol in ETF_SYMBOL_CANDIDATES[symbol]:
        try:
            log(f"抓取 {symbol}（{source}:{source_symbol}）...")
            if source == "yahoo":
                current, day_high, day_low = fetch_yahoo_price(source_symbol)
            else:
                current, day_high, day_low = fetch_stooq_price(source_symbol)
            return PriceResult(symbol, current, day_high, day_low, f"{source}:{source_symbol}", None)
        except Exception as exc:
            err = f"{source} 失败: {exc}"
            errors.append(err)
            log(f"{symbol} 抓取失败: {err}")

    return PriceResult(symbol, None, None, None, None, " | ".join(errors))


def _to_float(value: object) -> Optional[float]:
    """尽量将输入转换为 float，失败时返回 None。"""

    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _round_half_up(value: Optional[float], ndigits: int) -> Optional[float]:
    """按常规四舍五入（ROUND_HALF_UP）保留指定小数位。"""

    if value is None:
        return None
    quant = Decimal("1").scaleb(-ndigits)
    return float(Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP))


def _extract_fear_greed_from_payload(data: dict) -> Tuple[Optional[float], Optional[str]]:
    """兼容 CNN 不同返回结构，提取 Fear & Greed 值和描述。"""

    fg = data.get("fear_and_greed", {}) if isinstance(data, dict) else {}

    # 结构 A：fear_and_greed.now.value / valueText
    if isinstance(fg, dict):
        now_data = fg.get("now", {}) if isinstance(fg.get("now"), dict) else {}
        value = _to_float(now_data.get("value"))
        if value is None:
            value = _to_float(now_data.get("score"))
        if value is None:
            value = _to_float(fg.get("value"))
        if value is None:
            value = _to_float(fg.get("score"))
        rating = now_data.get("valueText") or now_data.get("rating") or fg.get("rating")
        if value is not None:
            return value, str(rating) if rating else None

    # 结构 B：历史数组里最后一个点（常见字段 y/value/score）
    history = data.get("fear_and_greed_historical")
    if isinstance(history, dict):
        items = history.get("data") or history.get("history") or history.get("values")
        if isinstance(items, list) and items:
            for item in reversed(items):
                if not isinstance(item, dict):
                    continue
                value = _to_float(item.get("y"))
                if value is None:
                    value = _to_float(item.get("value"))
                if value is None:
                    value = _to_float(item.get("score"))
                if value is not None:
                    rating = item.get("rating") or item.get("valueText")
                    return value, str(rating) if rating else None

    return None, None


def fetch_fear_greed() -> FearGreedResult:
    """抓取 CNN Fear & Greed 指标。

    失败时返回带 error 的结果，不抛异常，保证主流程可继续。
    """

    url = f"https://production.dataviz.cnn.io/index/fearandgreed/graphdata?ts={int(time.time() * 1000)}"
    cnn_headers = {
        "Referer": "https://www.cnn.com/markets/fear-and-greed",
        "Origin": "https://www.cnn.com",
    }
    try:
        log("抓取 CNN Fear & Greed...")
        data = http_get_json(url, extra_headers=cnn_headers)
        value, rating = _extract_fear_greed_from_payload(data)
        if value is None:
            raise RuntimeError(f"Fear & Greed 字段缺失，可用键: {list(data.keys())[:8] if isinstance(data, dict) else 'N/A'}")
        if rating:
            log(f"CNN Fear & Greed 抓取成功: {value:.1f} ({rating})")
        else:
            log(f"CNN Fear & Greed 抓取成功: {value:.1f}")
        return FearGreedResult(float(value), str(rating) if rating else None, None)
    except Exception as exc:
        error = f"CNN Fear & Greed 抓取失败: {exc}"
        log(error)
        return FearGreedResult(None, None, error)


def _extract_market_line(html: str, label: str) -> Tuple[Optional[str], Optional[str]]:
    """从 CNN 页面 HTML 中提取指定指数的点位和涨跌幅。"""

    # 先做实体和 \uXXXX 解码，提升在 script 内嵌 JSON 场景下的匹配概率
    normalized = html_lib.unescape(html)

    def _decode_unicode_escape(m: re.Match[str]) -> str:
        try:
            return chr(int(m.group(1), 16))
        except ValueError:
            return m.group(0)

    normalized = re.sub(r"\\u([0-9a-fA-F]{4})", _decode_unicode_escape, normalized)

    if label.upper() == "S&P 500":
        label_pattern = r"(?:S\s*&\s*P\s*500|S\s*P\s*500|SP500)"
    elif label.upper() == "NASDAQ":
        label_pattern = r"NASDAQ"
    else:
        label_pattern = re.escape(label)

    # 允许中间有标签/空白，兼容部分页面结构变化；百分比允许带/不带 %
    pattern = re.compile(
        rf"{label_pattern}(?:.|\n){{0,350}}?([0-9][0-9,]*(?:\.[0-9]+)?)(?:.|\n){{0,180}}?([-+]?\d+(?:\.[0-9]+)?%?)",
        re.IGNORECASE,
    )
    for m in pattern.finditer(normalized):
        value = _format_index_value(m.group(1).replace(",", ""))
        change = _format_percent(m.group(2))
        value_num = _to_float((value or "").replace(",", ""))
        change_num = _to_float((change or "").replace("%", ""))

        # 过滤误匹配：如把标签本身中的“500”或无意义 0 当作指数点位
        if value_num is None or change_num is None:
            continue
        if value_num < 1000 or value_num > 100000:
            continue
        if abs(change_num) > 20:
            continue
        return value, change

    return None, None


def _format_index_value(value: object) -> Optional[str]:
    """将指数点位格式化为带千分位字符串。"""

    num = _to_float(value)
    if num is None:
        return None
    return f"{num:,.2f}"


def _format_percent(value: object) -> Optional[str]:
    """将涨跌幅字段规范为 xx.xx%。"""

    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("%"):
            return text
        num = _to_float(text)
        if num is None:
            return None
        return f"{num:.2f}%"
    num = _to_float(value)
    if num is None:
        return None
    # 有些接口返回小数（0.0014 代表 0.14%）
    if abs(num) <= 1:
        num = num * 100
    return f"{num:.2f}%"


def _iter_dicts(node: object):
    """深度遍历 JSON 树中的所有 dict。"""

    if isinstance(node, dict):
        yield node
        for v in node.values():
            yield from _iter_dicts(v)
    elif isinstance(node, list):
        for item in node:
            yield from _iter_dicts(item)


def _extract_indices_from_embedded_json(html: str) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
    """从页面内嵌 JSON（如 __NEXT_DATA__）提取 S&P500/NASDAQ 数据。"""

    script_pattern = re.compile(
        r"<script[^>]*(?:id=['\"]__NEXT_DATA__['\"][^>]*)?[^>]*type=['\"]application/json['\"][^>]*>(.*?)</script>",
        re.IGNORECASE | re.DOTALL,
    )

    sp500_value = sp500_change = nasdaq_value = nasdaq_change = None

    for match in script_pattern.finditer(html):
        raw = html_lib.unescape(match.group(1).strip())
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue

        for d in _iter_dicts(payload):
            label_fields = [
                str(d.get("label", "")),
                str(d.get("name", "")),
                str(d.get("title", "")),
                str(d.get("symbol", "")),
                str(d.get("ticker", "")),
            ]
            label_text = " ".join(label_fields).upper().replace(" ", "")

            value = (
                d.get("value")
                if "value" in d
                else d.get("price", d.get("last", d.get("lastPrice", d.get("close"))))
            )
            change = d.get("changePercent", d.get("percentChange", d.get("pctChange", d.get("change_pct"))))

            value_text = _format_index_value(value)
            change_text = _format_percent(change)
            value_num = _to_float((value_text or "").replace(",", ""))
            change_num = _to_float((change_text or "").replace("%", ""))
            if value_num is None or change_num is None:
                continue
            if value_num < 1000 or value_num > 100000 or abs(change_num) > 20:
                continue

            if ("S&P500" in label_text or "SP500" in label_text) and (sp500_value is None or sp500_change is None):
                sp500_value = sp500_value or value_text
                sp500_change = sp500_change or change_text
            elif "NASDAQ" in label_text and (nasdaq_value is None or nasdaq_change is None):
                nasdaq_value = nasdaq_value or value_text
                nasdaq_change = nasdaq_change or change_text

            if sp500_value and sp500_change and nasdaq_value and nasdaq_change:
                return sp500_value, sp500_change, nasdaq_value, nasdaq_change

    return sp500_value, sp500_change, nasdaq_value, nasdaq_change


def _fetch_index_from_yahoo(symbol: str, label: str) -> IndexSnapshot:
    """从 Yahoo 抓取指数快照，作为 CNN 页面解析失败时的兜底。"""

    quote_url = f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={urlencode({'s': symbol})[2:]}"
    try:
        data = http_get_json(quote_url)
        quote_items = (((data.get("quoteResponse") or {}).get("result")) or []) if isinstance(data, dict) else []
        item = quote_items[0] if quote_items else {}
        value = _format_index_value(item.get("regularMarketPrice"))
        change = _format_percent(item.get("regularMarketChangePercent"))

        # 兼容仅返回涨跌额未返回百分比的场景
        if value and not change:
            change = _format_percent(item.get("regularMarketChange"))

        # 仍缺字段时，退回 chart 接口补齐
        if not value or not change:
            chart_url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=5d"
            chart = http_get_json(chart_url)
            result = (chart.get("chart") or {}).get("result") or []
            if result:
                info = result[0]
                meta = info.get("meta") or {}
                quote = (((info.get("indicators") or {}).get("quote") or [{}])[0])
                closes = [x for x in (quote.get("close") or []) if x is not None]
                last_close = float(closes[-1]) if closes else None
                prev_close = _to_float(meta.get("chartPreviousClose") or meta.get("previousClose"))
                if not value:
                    value = _format_index_value(meta.get("regularMarketPrice") or last_close)
                if not change:
                    pct = _to_float(meta.get("regularMarketChangePercent"))
                    if pct is None and last_close is not None and prev_close not in (None, 0):
                        pct = ((last_close - prev_close) / prev_close) * 100
                    change = _format_percent(pct)

        if not value or not change:
            raise RuntimeError("Yahoo 指数字段缺失")
        log(f"{label} 使用 Yahoo 兜底成功: {value} ({change})")
        return IndexSnapshot(value, change, None)
    except Exception as exc:
        return IndexSnapshot(None, None, f"{label} Yahoo 兜底失败: {exc}")


def _extract_index_from_any_json(data: object, label: str) -> IndexSnapshot:
    """从任意 JSON 结构中按标签提取指数点位和涨跌幅。"""

    target = label.upper().replace(" ", "")
    if target == "S&P500":
        aliases = {"S&P500", "SP500", "GSPC", "^GSPC"}
    elif target == "NASDAQ":
        aliases = {"NASDAQ", "IXIC", "^IXIC", "NASDAQCOMPOSITE"}
    else:
        aliases = {target}

    for d in _iter_dicts(data):
        label_fields = [
            str(d.get("label", "")),
            str(d.get("name", "")),
            str(d.get("title", "")),
            str(d.get("symbol", "")),
            str(d.get("ticker", "")),
            str(d.get("id", "")),
        ]
        label_text = "".join(label_fields).upper().replace(" ", "")
        if not any(alias in label_text for alias in aliases):
            continue

        value = d.get("value", d.get("price", d.get("last", d.get("lastPrice", d.get("close")))))
        change = d.get("changePercent", d.get("percentChange", d.get("pctChange", d.get("change_pct"))))
        value_text = _format_index_value(value)
        change_text = _format_percent(change)
        if value_text and change_text:
            return IndexSnapshot(value_text, change_text, None)

    return IndexSnapshot(None, None, f"{label} 未在 JSON 中解析到")


def _fetch_index_from_cnn_api(label: str) -> IndexSnapshot:
    """尝试 CNN 可能的市场接口（避免页面结构变化影响）。"""

    api_urls = [
        "https://production.dataviz.cnn.io/markets/indexes",
        "https://production.dataviz.cnn.io/markets/indexes?exchange=US",
        "https://production.dataviz.cnn.io/markets/overview",
    ]
    last_err: Optional[str] = None
    for url in api_urls:
        try:
            data = http_get_json(url, extra_headers={"Referer": "https://edition.cnn.com/markets/fear-and-greed"})
            snap = _extract_index_from_any_json(data, label)
            if snap.value and snap.change:
                log(f"{label} 使用 CNN API 兜底成功: {snap.value} ({snap.change})")
                return snap
            last_err = snap.error or f"{label} CNN API 未命中"
        except Exception as exc:
            last_err = f"{label} CNN API 请求失败: {exc}"
    return IndexSnapshot(None, None, last_err or f"{label} CNN API 兜底失败")


def _fetch_index_from_stooq(symbol: str, label: str) -> IndexSnapshot:
    """使用 Stooq 指数符号兜底。"""

    try:
        url = f"https://stooq.com/q/l/?s={symbol}&f=sd2t2ohlcv&h&e=csv"
        text = http_get_text(url)
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if len(lines) < 2:
            raise RuntimeError("Stooq 返回空数据")
        cols = [c.strip().lower() for c in lines[0].split(",")]
        vals = [v.strip() for v in lines[1].split(",")]
        row = dict(zip(cols, vals))
        close = _to_float(row.get("close"))
        open_ = _to_float(row.get("open"))
        if close is None:
            raise RuntimeError("Stooq close 缺失")
        change_pct: Optional[float] = None
        if open_ not in (None, 0):
            change_pct = ((close - open_) / open_) * 100
        value = _format_index_value(close)
        # Stooq 计算得到的是“百分比点”（如 0.3258 表示 0.3258%），
        # 转成比例后再交给 _format_percent（其会乘以 100）。
        change = _format_percent(change_pct / 100.0) if change_pct is not None else None
        if not value or not change:
            raise RuntimeError("Stooq 指数字段缺失")
        log(f"{label} 使用 Stooq 兜底成功: {value} ({change})")
        return IndexSnapshot(value, change, None)
    except Exception as exc:
        return IndexSnapshot(None, None, f"{label} Stooq 兜底失败: {exc}")


def _fetch_indices_from_cnn_dataviz() -> Tuple[IndexSnapshot, IndexSnapshot]:
    """从 CNN Dataviz 指数接口抓取 S&P500/NASDAQ（与页面红绿方向一致）。"""

    try:
        utc_date = datetime.now(ZoneInfo("UTC")).strftime("%Y-%m-%d")
        url = f"{CNN_DATAVIZ_API_BASE}/markets/index/{CNN_MARKET_INDEX_SYMBOLS}/{utc_date}"
        rows = http_get_json(
            url,
            extra_headers={"Referer": "https://edition.cnn.com/markets/fear-and-greed"},
        )
        if not isinstance(rows, list):
            raise RuntimeError(f"返回结构异常: {type(rows).__name__}")

        row_map: Dict[str, dict] = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            symbol = str(row.get("symbol", "")).strip()
            if symbol:
                row_map[symbol] = row

        def _to_snapshot(symbol: str, label: str) -> IndexSnapshot:
            row = row_map.get(symbol, {})
            value = _format_index_value(row.get("current_price"))
            # 该字段是比例值：-0.0043 -> -0.43%
            change = _format_percent(row.get("percent_change_from_prev_close"))
            if value and change:
                return IndexSnapshot(value, change, None)
            return IndexSnapshot(None, None, f"{label} Dataviz 字段缺失")

        sp500 = _to_snapshot("SP500-CME", "S&P 500")
        nasdaq = _to_snapshot("COMP-USA", "NASDAQ")
        return sp500, nasdaq
    except Exception as exc:
        err = f"CNN Dataviz 指数接口失败: {exc}"
        return IndexSnapshot(None, None, err), IndexSnapshot(None, None, err)


def fetch_cnn_market_snapshot() -> CnnMarketResult:
    """抓取 CNN Fear & Greed 页面上的 S&P 500 / NASDAQ 摘要。"""

    url = f"https://edition.cnn.com/markets/fear-and-greed?ts={int(time.time() * 1000)}"
    try:
        log("抓取 CNN Markets 页面（S&P 500 / NASDAQ）...")
        # 优先使用页面实际依赖的 Dataviz 指数接口，保证与红/绿展示一致。
        sp500_dv, nasdaq_dv = _fetch_indices_from_cnn_dataviz()
        sp500_value = sp500_dv.value
        sp500_change = sp500_dv.change
        nasdaq_value = nasdaq_dv.value
        nasdaq_change = nasdaq_dv.change

        html = None
        if (not sp500_value or not sp500_change) or (not nasdaq_value or not nasdaq_change):
            html = http_get_text(url)
            sp500_html_value, sp500_html_change = _extract_market_line(html, "S&P 500")
            nasdaq_html_value, nasdaq_html_change = _extract_market_line(html, "NASDAQ")
            sp500_value = sp500_value or sp500_html_value
            sp500_change = sp500_change or sp500_html_change
            nasdaq_value = nasdaq_value or nasdaq_html_value
            nasdaq_change = nasdaq_change or nasdaq_html_change

        # HTML 正则失败时，尝试解析页面内嵌 JSON
        if (not sp500_value or not sp500_change) or (not nasdaq_value or not nasdaq_change):
            if html is None:
                html = http_get_text(url)
            j_sp500_value, j_sp500_change, j_nasdaq_value, j_nasdaq_change = _extract_indices_from_embedded_json(html)
            sp500_value = sp500_value or j_sp500_value
            sp500_change = sp500_change or j_sp500_change
            nasdaq_value = nasdaq_value or j_nasdaq_value
            nasdaq_change = nasdaq_change or j_nasdaq_change

        # 若仍有缺失，先尝试 CNN 其它接口，再用 Yahoo/Stooq 兜底
        fallback_errors: List[str] = []
        if not sp500_value or not sp500_change:
            sp500_fb = _fetch_index_from_cnn_api("S&P 500")
            if not sp500_fb.value or not sp500_fb.change:
                sp500_fb = _fetch_index_from_yahoo("^GSPC", "S&P 500")
            if not sp500_fb.value or not sp500_fb.change:
                sp500_fb = _fetch_index_from_stooq("^spx", "S&P 500")
            sp500_value = sp500_value or sp500_fb.value
            sp500_change = sp500_change or sp500_fb.change
            if sp500_fb.error:
                fallback_errors.append(sp500_fb.error)

        if not nasdaq_value or not nasdaq_change:
            nasdaq_fb = _fetch_index_from_cnn_api("NASDAQ")
            if not nasdaq_fb.value or not nasdaq_fb.change:
                nasdaq_fb = _fetch_index_from_yahoo("^IXIC", "NASDAQ")
            if not nasdaq_fb.value or not nasdaq_fb.change:
                nasdaq_fb = _fetch_index_from_stooq("^ndq", "NASDAQ")
            nasdaq_value = nasdaq_value or nasdaq_fb.value
            nasdaq_change = nasdaq_change or nasdaq_fb.change
            if nasdaq_fb.error:
                fallback_errors.append(nasdaq_fb.error)

        if not sp500_value and not nasdaq_value:
            if fallback_errors:
                raise RuntimeError("页面中未匹配到 S&P 500/NASDAQ 数值，且兜底失败: " + " | ".join(fallback_errors))
            raise RuntimeError("页面中未匹配到 S&P 500/NASDAQ 数值")

        if sp500_value and sp500_change:
            log(f"CNN S&P 500: {sp500_value} ({sp500_change})")
        if nasdaq_value and nasdaq_change:
            log(f"CNN NASDAQ: {nasdaq_value} ({nasdaq_change})")

        return CnnMarketResult(
            sp500_value=sp500_value,
            sp500_change=sp500_change,
            nasdaq_value=nasdaq_value,
            nasdaq_change=nasdaq_change,
            error=None,
        )
    except Exception as exc:
        err = f"CNN Markets 页面抓取失败: {exc}"
        log(err)
        return CnnMarketResult(None, None, None, None, err)


def fetch_boc_usd_spot_sell() -> ExchangeRateResult:
    """抓取中行“美元”现汇卖出价。"""

    try:
        log("抓取中国银行美元现汇卖出价...")
        html = http_get_text(BOC_WHPJ_URL)
        parser = _BOCPriceTableParser(BOC_TARGET_CURRENCY)
        parser.feed(html)

        if not parser.target_row:
            raise RuntimeError("未匹配到美元行")
        if len(parser.target_row) < 8:
            raise RuntimeError(f"美元行字段不足: {parser.target_row}")

        # 列顺序: 货币名称, 现汇买入价, 现钞买入价, 现汇卖出价, 现钞卖出价, 中行折算价, 发布日期, 发布时间
        rate_text = parser.target_row[3]
        publish_raw = parser.target_row[6]
        publish_time = parser.target_row[7]
        rate = _to_float(rate_text)
        if rate is None:
            raise RuntimeError(f"现汇卖出价解析失败: {rate_text}")

        publish_date = publish_raw.split(" ")[0].replace("/", "-") if publish_raw else None
        log(f"中行美元现汇卖出价抓取成功: {rate:.4f}")
        return ExchangeRateResult(rate, publish_date, publish_time or None, None)
    except Exception as exc:
        err = f"中行汇率抓取失败: {exc}"
        log(err)
        return ExchangeRateResult(None, None, None, err)


def fetch_pingan_gold_price() -> GoldPriceResult:
    """抓取平安金活期当前金价（元/克）。"""

    try:
        log("抓取平安黄金价格...")
        data = http_get_json(
            PINGAN_GOLD_DETAIL_URL,
            extra_headers={"Referer": "https://b.pingan.com.cn/aum/mobile2/detail_gold_current.html"},
        )
        if not isinstance(data, dict):
            raise RuntimeError("平安黄金返回结构异常")
        if str(data.get("responseCode")) != "000000":
            raise RuntimeError(f"平安黄金接口返回非成功: {data.get('responseCode')}")

        payload = data.get("data") or {}
        price = _to_float(payload.get("price"))
        if price is None:
            raise RuntimeError("平安黄金 price 字段缺失")

        log(f"平安黄金价格抓取成功: {price:.2f} 元/克")
        return GoldPriceResult(price, None)
    except Exception as exc:
        err = f"平安黄金价格抓取失败: {exc}"
        log(err)
        return GoldPriceResult(None, err)


def fetch_jm_gold_fear_greed_index() -> GoldSentimentResult:
    """抓取 JM Bullion 黄金情绪指数。"""

    try:
        log("抓取 JM 黄金情绪指数...")
        url = f"{JM_GOLD_FEAR_GREED_URL}?v={datetime.now(BEIJING_TZ).strftime('%Y%m%d%H')}"
        data = http_get_json(url, extra_headers={"Referer": "https://www.jmbullion.com/fear-greed-index/"})
        if not isinstance(data, dict) or not data:
            raise RuntimeError("JM 黄金情绪指数返回为空")

        latest_date = max(str(k) for k in data.keys())
        value = _to_float(data.get(latest_date))
        if value is None:
            raise RuntimeError(f"JM 黄金情绪指数解析失败，日期: {latest_date}")

        log(f"JM 黄金情绪指数抓取成功: {value:.1f}")
        return GoldSentimentResult(value, None)
    except Exception as exc:
        err = f"JM 黄金情绪指数抓取失败: {exc}"
        log(err)
        return GoldSentimentResult(None, err)


def fetch_jpm_118uc_price() -> FundPriceResult:
    """抓取 118UC（JPM US Technology A (acc) - USD）当前净值。"""

    try:
        log("抓取 118UC 价格...")
        data = http_get_json(JPM_118UC_DATA_URL, extra_headers={"Referer": JPM_118UC_URL})
        if not isinstance(data, dict):
            raise RuntimeError("JPM 返回结构异常")

        fund_data = data.get("fundData") or {}
        share_class = fund_data.get("shareClass") or {}
        nav = share_class.get("nav") or {}
        value = _to_float(nav.get("price"))
        if value is None:
            raise RuntimeError("JPM NAV 价格字段缺失")

        log(f"118UC 价格抓取成功: {value:.2f}")
        return FundPriceResult(value, None)
    except Exception as exc:
        err = f"118UC 价格抓取失败: {exc}"
        log(err)
        return FundPriceResult(None, err)


def _fetch_blackrock_nav_price(url: str, code: str) -> FundPriceResult:
    """抓取 BlackRock 页面中的净值（美元）。"""

    try:
        log(f"抓取 {code} 价格...")
        html = http_get_text(url)
        pattern = re.compile(
            r'class="navAmount[^>]*>[\s\S]{0,1200}?class="header-nav-data"[^>]*>[\s\S]{0,40}?([0-9]+(?:\.[0-9]+)?)',
            re.IGNORECASE,
        )
        match = pattern.search(html)
        if not match:
            raise RuntimeError("BlackRock 净值字段未匹配到")

        value = _to_float(match.group(1))
        if value is None:
            raise RuntimeError(f"BlackRock 净值解析失败: {match.group(1)}")

        log(f"{code} 价格抓取成功: {value:.2f}")
        return FundPriceResult(value, None)
    except Exception as exc:
        err = f"{code} 价格抓取失败: {exc}"
        log(err)
        return FundPriceResult(None, err)


def fetch_blackrock_134uu_price() -> FundPriceResult:
    """抓取 134UU（贝莱德世界科技基金 A2 USD）当前净值。"""

    return _fetch_blackrock_nav_price(BLACKROCK_134UU_URL, "134UU")


def fetch_blackrock_077uc_price() -> FundPriceResult:
    """抓取 077UC（贝莱德智慧数据环球股票高息基金 A6 USD）当前净值。"""

    return _fetch_blackrock_nav_price(BLACKROCK_077UC_URL, "077UC")


def fetch_sc_172uc_price() -> FundPriceResult:
    """抓取 172UC（渣打页 LU2708338459）当前净值。"""

    try:
        log("抓取 172UC 价格...")
        query = (
            f'query MyQuery {{ morningStar(isin:"{SC_172UC_ISIN}",currency:"{SC_172UC_CURRENCY}") }}'
        )
        payload = {"query": query}
        data = http_post_json(SC_GRAPHQL_URL, payload, extra_headers={"Referer": SC_172UC_PAGE_URL})

        root = data.get("data") if isinstance(data, dict) else None
        raw_morningstar = root.get("morningStar") if isinstance(root, dict) else None
        if raw_morningstar is None:
            raise RuntimeError("SC morningStar 字段缺失")

        if isinstance(raw_morningstar, str):
            morningstar = json.loads(raw_morningstar)
        elif isinstance(raw_morningstar, dict):
            morningstar = raw_morningstar
        else:
            raise RuntimeError("SC morningStar 结构异常")

        performance = morningstar.get("performance") if isinstance(morningstar, dict) else None
        value = _to_float(performance.get("nav")) if isinstance(performance, dict) else None
        if value is None:
            raise RuntimeError("SC performance.nav 字段缺失")

        log(f"172UC 价格抓取成功: {value:.2f}")
        return FundPriceResult(value, None)
    except Exception as exc:
        err = f"172UC 价格抓取失败: {exc}"
        log(err)
        return FundPriceResult(None, err)


def _fetch_ab_nav_detail_price(share_class_id: str, code: str, page_url: str) -> FundPriceResult:
    """抓取 AB 基金 nav-detail 中的 dailyNAV。"""

    try:
        log(f"抓取 {code} 价格...")
        url = f"{AB_FUNDS_API_BASE_URL}/v2/funds/hk/zh-hk/investor/{share_class_id}/nav-detail"
        data = http_get_json(url, extra_headers={"Referer": page_url})
        if not isinstance(data, dict):
            raise RuntimeError("AB nav-detail 返回结构异常")

        value = _to_float(data.get("dailyNAV"))
        if value is None:
            raise RuntimeError("AB dailyNAV 字段缺失")

        log(f"{code} 价格抓取成功: {value:.2f}")
        return FundPriceResult(value, None)
    except Exception as exc:
        err = f"{code} 价格抓取失败: {exc}"
        log(err)
        return FundPriceResult(None, err)


def fetch_ab_089uu_price() -> FundPriceResult:
    """抓取 089UU（AB 低波幅策略股票基金）当前净值。"""

    return _fetch_ab_nav_detail_price(AB_089UU_SHARE_CLASS_ID, "089UU", AB_089UU_PAGE_URL)


def fetch_ab_123uu_price() -> FundPriceResult:
    """抓取 123UU（AB 低波幅股票总回报基金）当前净值。"""

    return _fetch_ab_nav_detail_price(AB_123UU_SHARE_CLASS_ID, "123UU", AB_123UU_PAGE_URL)


def configure_log_xlsx_path(logfile: Optional[str]) -> None:
    """Configure log xlsx path; default uses LOG_XLSX_FILENAME."""

    global LOG_XLSX_PATH_OVERRIDE
    if logfile:
        LOG_XLSX_PATH_OVERRIDE = Path(logfile).expanduser()
    else:
        LOG_XLSX_PATH_OVERRIDE = None


def _log_xlsx_path() -> Path:
    """Return the log Excel file path."""

    return LOG_XLSX_PATH_OVERRIDE or (Path(__file__).resolve().parent / LOG_XLSX_FILENAME)


def _parse_index_value(value: Optional[str]) -> Optional[float]:
    """将形如 '6,878.88' 的指数文本转换为 float。"""

    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    return _to_float(text)


def _parse_percent_ratio(value: Optional[str]) -> Optional[float]:
    """将百分比文本转换为 Excel 百分比底层值（如 '23.34%' -> 0.2334）。"""

    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    if text.endswith("%"):
        pct = _to_float(text[:-1].strip())
        return None if pct is None else pct / 100.0

    num = _to_float(text)
    if num is None:
        return None
    return num / 100.0 if abs(num) > 1 else num


def append_market_snapshot_to_xlsx(
    snapshot_time: datetime,
    etf_prices: List[PriceResult],
    fx_rate: Optional[float],
    fg_value: Optional[float],
    sp500_value: Optional[str],
    sp500_change: Optional[str],
    nasdaq_value: Optional[str],
    nasdaq_change: Optional[str],
    gold_price: Optional[float],
    gold_fear_greed: Optional[float],
    price_118uc: Optional[float],
    price_134uu: Optional[float],
    price_077uc: Optional[float],
    price_172uc: Optional[float],
    price_089uu: Optional[float],
    price_123uu: Optional[float],
) -> None:
    """将行情快照写入 xlsx，最新记录插入到第 2 行。"""

    try:
        from openpyxl import Workbook, load_workbook

        path = _log_xlsx_path()
        path.parent.mkdir(parents=True, exist_ok=True)
        if not path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = LOG_XLSX_SHEET1_NAME
            ws.append(LOG_XLSX_HEADERS)
            wb.save(path)
            log(f"未找到 {LOG_XLSX_FILENAME}，已自动创建")

        wb = load_workbook(path)
        ws = wb[LOG_XLSX_SHEET1_NAME] if LOG_XLSX_SHEET1_NAME in wb.sheetnames else wb.create_sheet(LOG_XLSX_SHEET1_NAME, 0)

        if ws.max_row < 1:
            ws.append(LOG_XLSX_HEADERS)
        else:
            for col_idx, header in enumerate(LOG_XLSX_HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        ws.insert_rows(2)
        excel_time = snapshot_time.replace(tzinfo=None, second=0, microsecond=0)
        ws.cell(row=2, column=1, value=excel_time)
        ws.cell(row=2, column=2, value=fx_rate)
        ws.cell(row=2, column=3, value=_round_half_up(fg_value, 1))
        ws.cell(row=2, column=4, value=_parse_index_value(sp500_value))
        ws.cell(row=2, column=5, value=_parse_percent_ratio(sp500_change))
        ws.cell(row=2, column=6, value=_parse_index_value(nasdaq_value))
        ws.cell(row=2, column=7, value=_parse_percent_ratio(nasdaq_change))
        ws.cell(row=2, column=8, value=gold_price)
        ws.cell(row=2, column=9, value=gold_fear_greed)
        ws.cell(row=2, column=10, value=price_118uc)
        ws.cell(row=2, column=11, value=price_134uu)
        ws.cell(row=2, column=12, value=price_077uc)
        ws.cell(row=2, column=13, value=price_172uc)
        ws.cell(row=2, column=14, value=price_089uu)
        ws.cell(row=2, column=15, value=price_123uu)

        ws.cell(row=2, column=1).number_format = "yyyy/m/d h:mm"
        ws.cell(row=2, column=4).number_format = "#,##0.00"
        ws.cell(row=2, column=5).number_format = "0.00%"
        ws.cell(row=2, column=6).number_format = "#,##0.00"
        ws.cell(row=2, column=7).number_format = "0.00%"
        ws.cell(row=2, column=8).number_format = "#,##0.00"
        ws.cell(row=2, column=9).number_format = "0.0"
        ws.cell(row=2, column=10).number_format = "#,##0.00"
        ws.cell(row=2, column=11).number_format = "#,##0.00"
        ws.cell(row=2, column=12).number_format = "#,##0.00"
        ws.cell(row=2, column=13).number_format = "#,##0.00"
        ws.cell(row=2, column=14).number_format = "#,##0.00"
        ws.cell(row=2, column=15).number_format = "#,##0.00"

        ws_etf = wb[LOG_XLSX_SHEET2_NAME] if LOG_XLSX_SHEET2_NAME in wb.sheetnames else wb.create_sheet(LOG_XLSX_SHEET2_NAME)
        if ws_etf.max_row < 1:
            ws_etf.append(LOG_XLSX_ETF_HEADERS)
        else:
            for col_idx, header in enumerate(LOG_XLSX_ETF_HEADERS, start=1):
                ws_etf.cell(row=1, column=col_idx, value=header)

        ws_etf.insert_rows(2)
        ws_etf.cell(row=2, column=1, value=excel_time)
        ws_etf.cell(row=2, column=1).number_format = "yyyy/m/d h:mm"

        etf_price_map = {p.symbol: p for p in etf_prices}
        for idx, symbol in enumerate(LOG_XLSX_ETF_SYMBOL_ORDER):
            p = etf_price_map.get(symbol)
            base_col = 2 + idx * 3
            ws_etf.cell(row=2, column=base_col, value=p.current if p else None)
            ws_etf.cell(row=2, column=base_col + 1, value=p.day_low if p else None)
            ws_etf.cell(row=2, column=base_col + 2, value=p.day_high if p else None)
            ws_etf.cell(row=2, column=base_col).number_format = "#,##0.0000"
            ws_etf.cell(row=2, column=base_col + 1).number_format = "#,##0.0000"
            ws_etf.cell(row=2, column=base_col + 2).number_format = "#,##0.0000"

        wb.save(path)
    except ImportError:
        log("未安装 openpyxl，跳过写入 xlsx")
    except Exception as exc:
        log(f"写入 {LOG_XLSX_FILENAME} 失败: {exc}")


def _format_xlsx_cell_text(value: object, as_percent: bool = False) -> str:
    """将 xlsx 单元格值转换为适合邮件展示的文本。"""

    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric_value = float(value)
        if as_percent:
            return f"{numeric_value * 100:.2f}%"
        if numeric_value.is_integer():
            return str(int(numeric_value))
        return f"{numeric_value:.4f}".rstrip("0").rstrip(".")
    return str(value).strip() or "-"


def read_sheet3_report_rows() -> Tuple[List[List[str]], Optional[str]]:
    """读取 Sheet3 的 A/B/C/E/G/I/K/M/O/Q 列第 1~3 行。"""

    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "未安装 openpyxl"

    path = _log_xlsx_path()
    if not path.exists():
        return [], f"log file not found: {path}"

    columns = ("A", "B", "C", "E", "G", "I", "K", "M", "O", "Q")
    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        if LOG_XLSX_SHEET3_NAME not in wb.sheetnames:
            return [], f"未找到工作表 {LOG_XLSX_SHEET3_NAME}"

        ws = wb[LOG_XLSX_SHEET3_NAME]
        header_values = [str(ws[f"{col}1"].value or "").upper() for col in columns]
        rate_columns = ["RAT" in header for header in header_values]
        rows: List[List[str]] = []
        for row_idx in range(1, 4):
            row_values = [
                _format_xlsx_cell_text(
                    ws[f"{col}{row_idx}"].value,
                    as_percent=(row_idx > 1 and rate_columns[col_idx]),
                )
                for col_idx, col in enumerate(columns)
            ]
            rows.append(row_values)
        return rows, None
    except Exception as exc:
        return [], f"读取 {LOG_XLSX_SHEET3_NAME} 失败: {exc}"
    finally:
        if wb is not None:
            wb.close()


def _format_aligned_rows(rows: List[List[str]]) -> List[str]:
    """按列宽对齐二维文本。"""

    if not rows:
        return []

    col_count = len(rows[0])
    widths = [0] * col_count
    for row in rows:
        for col_idx in range(col_count):
            widths[col_idx] = max(widths[col_idx], len(row[col_idx]))

    formatted_lines: List[str] = []
    for row in rows:
        line = "  ".join(f"{row[col_idx]:<{widths[col_idx]}}" for col_idx in range(col_count)).rstrip()
        formatted_lines.append(line)
    return formatted_lines


def read_config(path: str) -> SmtpConfig:
    """读取并校验 SMTP 配置。

    配置缺失属于“不可恢复错误”，由 main() 捕获后直接退出。
    """

    config_path = Path(path)
    if not config_path.is_file():
        raise FileNotFoundError(f"无法读取配置文件: {path}")

    parser = configparser.ConfigParser()
    encodings_to_try = ("utf-8-sig", "utf-8", "gb18030", "gbk")
    decode_failed = True
    for encoding in encodings_to_try:
        try:
            with config_path.open("r", encoding=encoding) as fp:
                parser.read_file(fp)
            decode_failed = False
            break
        except UnicodeDecodeError:
            parser.clear()
    if decode_failed:
        tried = ", ".join(encodings_to_try)
        raise ValueError(
            f"配置文件编码无法解码（已尝试: {tried}）。请将配置文件保存为 UTF-8 或 GB18030: {path}"
        )
    if "smtp" not in parser:
        raise KeyError("配置缺少 [smtp] 段")

    section = parser["smtp"]
    required = ["host", "port", "username", "password"]
    missing = [key for key in required if not section.get(key)]
    if missing:
        raise KeyError(f"配置缺少字段: {', '.join(missing)}")

    port = section.getint("port")
    from_email = section.get("from_email", section["username"])
    use_tls = section.getboolean("use_tls", fallback=(port == 587))
    use_ssl = section.getboolean("use_ssl", fallback=(port == 465))
    if use_tls and use_ssl:
        raise ValueError("配置冲突: use_tls 与 use_ssl 不能同时为 true")

    return SmtpConfig(
        host=section["host"],
        port=port,
        username=section["username"],
        password=section["password"],
        from_email=from_email,
        use_tls=use_tls,
        use_ssl=use_ssl,
    )


def build_report() -> Tuple[str, str, str]:
    """构建邮件主题和正文。

    规则：若任一关键数据缺失（任一 ETF 或 Fear&Greed），主题标记为 NOK。
    """

    prices = [fetch_price(symbol) for symbol in ETF_SYMBOL_CANDIDATES]
    fg = fetch_fear_greed()
    market = fetch_cnn_market_snapshot()
    fx = fetch_boc_usd_spot_sell()
    gold_price = fetch_pingan_gold_price()
    gold_fear_greed = fetch_jm_gold_fear_greed_index()
    price_118uc = fetch_jpm_118uc_price()
    price_134uu = fetch_blackrock_134uu_price()
    price_077uc = fetch_blackrock_077uc_price()
    price_172uc = fetch_sc_172uc_price()
    price_089uu = fetch_ab_089uu_price()
    price_123uu = fetch_ab_123uu_price()

    subject = "USSER 美股ETF交易提醒"

    report_time = datetime.now(BEIJING_TZ)
    line_time_prefix = report_time.strftime("%Y-%m-%d %H:%M")

    append_market_snapshot_to_xlsx(
        snapshot_time=report_time,
        etf_prices=prices,
        fx_rate=fx.rate,
        fg_value=fg.value,
        sp500_value=market.sp500_value,
        sp500_change=market.sp500_change,
        nasdaq_value=market.nasdaq_value,
        nasdaq_change=market.nasdaq_change,
        gold_price=gold_price.value,
        gold_fear_greed=gold_fear_greed.value,
        price_118uc=price_118uc.value,
        price_134uu=price_134uu.value,
        price_077uc=price_077uc.value,
        price_172uc=price_172uc.value,
        price_089uu=price_089uu.value,
        price_123uu=price_123uu.value,
    )

    sheet3_rows, sheet3_error = read_sheet3_report_rows()
    log("Sheet3（A/B/C/E/G/I/K/M/O/Q，1-3行）:")
    if sheet3_error:
        log(f"  数据缺失: {sheet3_error}")
    else:
        for row in _format_aligned_rows(sheet3_rows):
            log(f"  {row}")

    lines = []

    lines.append(f"{'类别':<10}{'时间':<18}{'当前价':<14}{'当日最低':<14}{'当日最高':<14}")
    for p in prices:
        if p.current is None or p.day_high is None or p.day_low is None:
            lines.append(f"{p.symbol:<10}{line_time_prefix:<18}{'数据缺失':<14}{'-':<14}{'-':<14}")
            if p.error:
                lines.append(f"  说明: {p.error}")
        else:
            lines.append(
                f"{p.symbol:<10}{line_time_prefix:<18}{p.current:<14.4f}{p.day_low:<14.4f}{p.day_high:<14.4f}"
            )

    lines.append("")
    lines.append("Sheet3（A/B/C/E/G/I/K/M/O/Q，1-3行）:")
    if sheet3_error:
        lines.append(f"数据缺失: {sheet3_error}")
    else:
        lines.extend(_format_aligned_rows(sheet3_rows))

    lines.append("Fear & Greed:")
    if fg.value is None:
        lines.append(f"{line_time_prefix} - 数据缺失，原因: {fg.error}")
    else:
        rating_text = f" ({fg.rating})" if fg.rating else ""
        lines.append(f"{line_time_prefix} - {fg.value:.1f}{rating_text}")

    lines.append("CNN Markets:")
    if market.error:
        lines.append(f"{line_time_prefix} - 数据缺失，原因: {market.error}")
    else:
        if market.sp500_value and market.sp500_change:
            lines.append(f"{line_time_prefix} - S&P 500: {market.sp500_value} {market.sp500_change}")
        else:
            lines.append(f"{line_time_prefix} - S&P 500: 数据缺失")
        if market.nasdaq_value and market.nasdaq_change:
            lines.append(f"{line_time_prefix} - NASDAQ: {market.nasdaq_value} {market.nasdaq_change}")
        else:
            lines.append(f"{line_time_prefix} - NASDAQ: 数据缺失")

    lines.append("")
    lines.append("美元汇率：")
    if fx.rate is None:
        lines.append(f"{line_time_prefix:<20}{'数据缺失':<12}")
        if fx.error:
            lines.append(f"  说明: {fx.error}")
    else:
        lines.append(f"{line_time_prefix:<20}{fx.rate:<12.4f}")
        if fx.publish_date and fx.publish_time:
            lines.append(f"  牌价发布时间: {fx.publish_date} {fx.publish_time}")

    lines.append("")
    lines.append("金价 & 指数：")
    if gold_price.value is None or gold_fear_greed.value is None:
        price_text = f"{gold_price.value:.2f}" if gold_price.value is not None else "数据缺失"
        fear_greed_text = f"{gold_fear_greed.value:.1f}" if gold_fear_greed.value is not None else "数据缺失"
        lines.append(f"{line_time_prefix:<20}{price_text:<12}{fear_greed_text:<12}")
        if gold_price.error:
            lines.append(f"  黄金价格说明: {gold_price.error}")
        if gold_fear_greed.error:
            lines.append(f"  黄金情绪说明: {gold_fear_greed.error}")
    else:
        lines.append(f"{line_time_prefix:<20}{gold_price.value:<12.2f}{gold_fear_greed.value:<12.1f}")

    body_text = "\n".join(lines)

    html_parts = [
        "<html><body style=\"font-family:'Microsoft YaHei',Arial,sans-serif;font-size:14px;color:#111;\">",
        "<table style='border-collapse:collapse;margin-bottom:0;'>",
        "<thead><tr>",
        "<th style='padding:4px 10px 4px 0;text-align:left;'>类别</th>",
        "<th style='padding:4px 10px 4px 0;text-align:left;'>时间</th>",
        "<th style='padding:4px 10px 4px 0;text-align:left;'>当前价</th>",
        "<th style='padding:4px 10px 4px 0;text-align:left;'>当日最低</th>",
        "<th style='padding:4px 10px 4px 0;text-align:left;'>当日最高</th>",
        "</tr></thead><tbody>",
    ]

    for p in prices:
        if p.current is None or p.day_high is None or p.day_low is None:
            current_text = "数据缺失"
            high_text = "-"
            low_text = "-"
        else:
            current_text = f"{p.current:.4f}"
            high_text = f"{p.day_high:.4f}"
            low_text = f"{p.day_low:.4f}"
        html_parts.append(
            "<tr>"
            f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(p.symbol)}</td>"
            f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(line_time_prefix)}</td>"
            f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(current_text)}</td>"
            f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(low_text)}</td>"
            f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(high_text)}</td>"
            "</tr>"
        )
        if p.error:
            html_parts.append(
                "<tr><td colspan='5' style='padding:2px 0 2px 0;color:#666;'>"
                f"{html_lib.escape(p.symbol)} 说明: {html_lib.escape(p.error)}"
                "</td></tr>"
            )
    html_parts.append("</tbody></table>")

    html_parts.append("<table style='border-collapse:collapse;margin:4px 0 2px 0;'><tbody>")
    if sheet3_error:
        html_parts.append(
            "<tr><td style='padding:2px 10px 2px 0;color:#666;'>"
            f"Sheet3（A/B/C/E/G/I/K/M/O/Q，1-3行）数据缺失: {html_lib.escape(sheet3_error)}"
            "</td></tr>"
        )
    else:
        for row_idx, row_values in enumerate(sheet3_rows):
            cell_style = "font-weight:700;color:#000;" if row_idx == 0 else "color:#111;"
            cells = "".join(
                f"<td style='padding:2px 16px 2px 0;text-align:left;white-space:nowrap;{cell_style}'>"
                f"{html_lib.escape(value)}</td>"
                for value in row_values
            )
            html_parts.append(f"<tr>{cells}</tr>")
    html_parts.append("</tbody></table>")

    html_parts.append("<div style='margin-top:0;font-weight:700;'>Fear &amp; Greed:</div>")
    if fg.value is None:
        html_parts.append(
            f"<div>{html_lib.escape(line_time_prefix)} - 数据缺失，原因: {html_lib.escape(str(fg.error))}</div>"
        )
    else:
        rating_text = f" ({fg.rating})" if fg.rating else ""
        html_parts.append(
            f"<div>{html_lib.escape(line_time_prefix)} - {fg.value:.1f}{html_lib.escape(rating_text)}</div>"
        )

    html_parts.append("<div style='margin-top:0;font-weight:700;'>CNN Markets:</div>")
    if market.error:
        html_parts.append(
            f"<div>{html_lib.escape(line_time_prefix)} - 数据缺失，原因: {html_lib.escape(str(market.error))}</div>"
        )
    else:
        if market.sp500_value and market.sp500_change:
            html_parts.append(
                f"<div>{html_lib.escape(line_time_prefix)} - S&amp;P 500: "
                f"{html_lib.escape(market.sp500_value)} "
                f"<span style='font-weight:700;color:#000;'>{html_lib.escape(market.sp500_change)}</span></div>"
            )
        else:
            html_parts.append(f"<div>{html_lib.escape(line_time_prefix)} - S&amp;P 500: 数据缺失</div>")
        if market.nasdaq_value and market.nasdaq_change:
            html_parts.append(
                f"<div>{html_lib.escape(line_time_prefix)} - NASDAQ: "
                f"{html_lib.escape(market.nasdaq_value)} "
                f"<span style='font-weight:700;color:#000;'>{html_lib.escape(market.nasdaq_change)}</span></div>"
            )
        else:
            html_parts.append(f"<div>{html_lib.escape(line_time_prefix)} - NASDAQ: 数据缺失</div>")

    html_parts.append("<div style='margin-top:0;font-weight:700;'>美元汇率：</div>")
    html_parts.append("<table style='border-collapse:collapse;margin-bottom:4px;'><tbody>")
    if fx.rate is None:
        html_parts.append(
            f"<tr><td style='padding:2px 10px 2px 0;'>{html_lib.escape(line_time_prefix)}</td>"
            "<td style='padding:2px 10px 2px 0;'>数据缺失</td></tr>"
        )
        if fx.error:
            html_parts.append(f"<tr><td colspan='2' style='color:#666;'>说明: {html_lib.escape(fx.error)}</td></tr>")
    else:
        html_parts.append(
            f"<tr><td style='padding:2px 10px 2px 0;'>{html_lib.escape(line_time_prefix)}</td>"
            f"<td style='padding:2px 10px 2px 0;'>{fx.rate:.4f}</td></tr>"
        )
        if fx.publish_date and fx.publish_time:
            html_parts.append(
                "<tr><td colspan='2' style='color:#666;'>"
                f"牌价发布时间: {html_lib.escape(fx.publish_date)} {html_lib.escape(fx.publish_time)}"
                "</td></tr>"
            )
    html_parts.append("</tbody></table>")

    html_parts.append("<div style='margin-top:0;font-weight:700;'>金价 &amp; 指数：</div>")
    html_parts.append("<table style='border-collapse:collapse;'><tbody>")
    price_text = f"{gold_price.value:.2f}" if gold_price.value is not None else "数据缺失"
    fear_greed_text = f"{gold_fear_greed.value:.1f}" if gold_fear_greed.value is not None else "数据缺失"
    html_parts.append(
        f"<tr><td style='padding:2px 10px 2px 0;'>{html_lib.escape(line_time_prefix)}</td>"
        f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(price_text)}</td>"
        f"<td style='padding:2px 10px 2px 0;'>{html_lib.escape(fear_greed_text)}</td></tr>"
    )
    if gold_price.error:
        html_parts.append(
            "<tr><td colspan='3' style='color:#666;'>"
            f"黄金价格说明: {html_lib.escape(gold_price.error)}"
            "</td></tr>"
        )
    if gold_fear_greed.error:
        html_parts.append(
            "<tr><td colspan='3' style='color:#666;'>"
            f"黄金情绪说明: {html_lib.escape(gold_fear_greed.error)}"
            "</td></tr>"
        )
    html_parts.append("</tbody></table>")

    html_parts.append("</body></html>")
    body_html = "".join(html_parts)
    return subject, body_text, body_html


def send_email(cfg: SmtpConfig, to_email: str, subject: str, body_text: str, body_html: str) -> None:
    """通过 SMTP 发送邮件。"""

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = cfg.from_email
    msg["To"] = to_email
    msg.attach(MIMEText(body_text, _subtype="plain", _charset="utf-8"))
    msg.attach(MIMEText(body_html, _subtype="html", _charset="utf-8"))

    log(f"发送邮件到 {to_email}...")
    smtp_cls = smtplib.SMTP_SSL if cfg.use_ssl else smtplib.SMTP
    with smtp_cls(cfg.host, cfg.port, timeout=15) as server:
        if cfg.use_tls:
            # 587 端口常见为 STARTTLS 升级握手
            server.ehlo()
            server.starttls()
            server.ehlo()
        auth_caps = server.esmtp_features.get("auth", "")
        auth_methods = {method.strip().upper() for method in auth_caps.split()}
        if "LOGIN" in auth_methods:
            # 优先使用 AUTH LOGIN，便于拿到服务端返回的明确认证失败原因（如 535）
            code, resp = server.docmd("AUTH", "LOGIN")
            if code != 334:
                raise smtplib.SMTPAuthenticationError(code, resp)
            user_b64 = base64.b64encode(cfg.username.encode("utf-8")).decode("ascii")
            code, resp = server.docmd(user_b64)
            if code != 334:
                raise smtplib.SMTPAuthenticationError(code, resp)
            pwd_b64 = base64.b64encode(cfg.password.encode("utf-8")).decode("ascii")
            code, resp = server.docmd(pwd_b64)
            if code != 235:
                raise smtplib.SMTPAuthenticationError(code, resp)
        else:
            server.login(cfg.username, cfg.password)
        server.sendmail(cfg.from_email, [to_email], msg.as_string())
    log("邮件发送完成")


def is_weekday_beijing() -> bool:
    """是否为北京时间周一到周五。"""

    return datetime.now(BEIJING_TZ).weekday() < 5


def print_email_preview(subject: str, body: str) -> None:
    """在控制台打印即将发送的邮件内容。"""

    log("即将发送的邮件内容如下：")
    print("-" * 60)
    print(f"Subject: {subject}")
    print(body)
    print("-" * 60)


def run_once(cfg: Optional[SmtpConfig], email_to: Optional[str]) -> int:
    """执行一次抓取+发信。

    发信失败时打印邮件内容预览，便于在无 SMTP 环境下排查。
    """

    subject, body_text, body_html = build_report()
    # 按需求：发邮件前，先把邮件内容打印给用户
    print_email_preview(subject, body_text)

    # 按需求：若未提供 --config 或 --emailto，则跳过发信流程
    if cfg is None or not email_to:
        log("未提供 --config 与 --emailto 的完整参数，跳过发邮件环节")
        return 0

    try:
        send_email(cfg, email_to, subject, body_text, body_html)
        return 0
    except smtplib.SMTPAuthenticationError as exc:
        reason = (
            exc.smtp_error.decode("utf-8", errors="replace")
            if isinstance(exc.smtp_error, bytes)
            else str(exc.smtp_error)
        )
        log(f"邮件认证失败 (SMTP {exc.smtp_code}): {reason}")
        return 1
    except Exception as exc:
        log(f"邮件发送失败: {exc}")
        return 1


def schedule_loop(remind_hhmm: str, cfg: Optional[SmtpConfig], email_to: Optional[str]) -> int:
    """定时循环：每天到指定 HH:MM 时触发一次（仅工作日）。"""

    log(f"进入定时模式，提醒时间(北京): {remind_hhmm}，仅周一至周五触发")
    # 记录上次触发日期，防止同一天重复触发
    last_trigger_date: Optional[str] = None

    while True:
        now = datetime.now(BEIJING_TZ)
        date_key = now.strftime("%Y-%m-%d")
        now_hhmm = now.strftime("%H:%M")

        if now_hhmm == remind_hhmm and date_key != last_trigger_date:
            if is_weekday_beijing():
                log("到达提醒时间，开始执行...")
                run_once(cfg, email_to)
                last_trigger_date = date_key
            else:
                log("到达提醒时间但今天是周末，跳过")
                last_trigger_date = date_key

        # 轮询间隔 30 秒，降低 CPU 占用
        time.sleep(30)


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    # 使用 conflict_handler='resolve'，即使文件被误合并导致重复定义同名参数，
    # 也不会在 --help 阶段直接崩溃（后定义会覆盖先定义）。
    parser = argparse.ArgumentParser(description="ETF 邮件提醒脚本", conflict_handler="resolve")

    def add_or_replace_argument(*names: str, **kwargs: object) -> None:
        """防御性添加参数：若同名参数已存在则先移除再添加。"""

        option_names = {name for name in names if isinstance(name, str) and name.startswith("-")}
        if option_names:
            for action in list(parser._actions):
                if option_names.intersection(set(action.option_strings)):
                    parser._remove_action(action)
                    for option in action.option_strings:
                        parser._option_string_actions.pop(option, None)
        parser.add_argument(*names, **kwargs)

    add_or_replace_argument("--remind", default="09:30", help="北京时间提醒时间，格式 HH:MM")
    add_or_replace_argument("--emailto", help="收件人邮箱；不传则跳过发邮件")
    add_or_replace_argument("--config", help="SMTP 配置文件路径；不传则跳过发邮件")
    add_or_replace_argument("--once", action="store_true", help="立即执行一次并退出")
    add_or_replace_argument("--logfile", help=r"Path to log xlsx, e.g. d:\sync\usser_log.xlsx")
    return parser.parse_args()


def validate_hhmm(value: str) -> None:
    """校验 --remind 时间格式必须为 HH:MM。"""

    try:
        time.strptime(value, "%H:%M")
    except ValueError as exc:
        raise ValueError("--remind 必须是 HH:MM 格式") from exc


def main() -> int:
    """程序入口。"""

    args = parse_args()
    validate_hhmm(args.remind)
    configure_log_xlsx_path(args.logfile)

    cfg: Optional[SmtpConfig] = None
    if args.config and args.emailto:
        try:
            cfg = read_config(args.config)
        except Exception as exc:
            # 配置错误是不可恢复问题：直接退出并返回非 0
            log(f"配置错误: {exc}")
            return 2
    else:
        log("未传入完整的 --config 和 --emailto，程序将仅打印报告并跳过发邮件")

    if args.once:
        return run_once(cfg, args.emailto)
    return schedule_loop(args.remind, cfg, args.emailto)


if __name__ == "__main__":
    sys.exit(main())
